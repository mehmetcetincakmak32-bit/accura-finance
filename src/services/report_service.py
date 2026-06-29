"""
Accura Finance - Rapor Servisi
PDF ve Excel rapor oluşturma işlemleri
"""

import os
import logging
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from io import BytesIO

from src.utils.logger import setup_logger

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image, PageBreak,
                                     KeepTogether)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportService:
    """Rapor oluşturma servisi"""

    def __init__(self):
        self.logger = setup_logger('ReportService')
        self._ensure_fonts()

    def _ensure_fonts(self):
        self._font_registered = False
        if HAS_REPORTLAB:
            font_paths = [
                'C:\\Windows\\Fonts\\arial.ttf',
                'C:\\Windows\\Fonts\\Arial.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            ]
            for fp in font_paths:
                if os.path.exists(fp):
                    try:
                        pdfmetrics.registerFont(TTFont('CustomFont', fp))
                        self._font_registered = True
                        break
                    except Exception:
                        continue

    def _get_font_name(self):
        return 'CustomFont' if self._font_registered else 'Helvetica'

    def generate_pdf_report(self, title: str, data: List[Dict],
                            columns: List[Dict], filename: str) -> str:
        if not HAS_REPORTLAB:
            self.logger.error("reportlab kurulu degil. PDF olusturulamadi.")
            return ''

        try:
            output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'reports'
            )
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)

            doc = SimpleDocTemplate(
                filepath, pagesize=A4,
                rightMargin=15*mm, leftMargin=15*mm,
                topMargin=15*mm, bottomMargin=15*mm
            )

            styles = getSampleStyleSheet()
            font_name = self._get_font_name()
            story = []

            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Title'],
                fontName=font_name, fontSize=16, spaceAfter=20,
                alignment=1
            )
            story.append(Paragraph(title, title_style))
            story.append(Paragraph(
                f"Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                ParagraphStyle('DateStyle', fontName=font_name, fontSize=8,
                              textColor=colors.grey, alignment=1)
            ))
            story.append(Spacer(1, 10*mm))

            headers = [col.get('label', col.get('field', '')) for col in columns]
            col_widths = [col.get('width', 60) * mm for col in columns]

            table_data = [headers]
            for row in data:
                table_row = []
                for col in columns:
                    field = col.get('field', '')
                    val = row.get(field, '')
                    if isinstance(val, float):
                        val = f"{val:,.2f}"
                    elif val is None:
                        val = ''
                    table_row.append(str(val))
                table_data.append(table_row)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F3F4')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(table)
            doc.build(story)

            self.logger.info(f"PDF rapor olusturuldu: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"PDF rapor olusturma hatasi: {e}")
            return ''

    def generate_excel_report(self, title: str, data: List[Dict],
                              columns: List[Dict], filename: str) -> str:
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl kurulu degil. Excel olusturulamadi.")
            return ''

        try:
            output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'reports'
            )
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31]

            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = [col.get('label', col.get('field', '')) for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, col in enumerate(columns, 1):
                    field = col.get('field', '')
                    val = row_data.get(field, '')
                    if isinstance(val, float):
                        val = round(val, 2)
                    elif val is None:
                        val = ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.00'

            for col_idx, col in enumerate(columns, 1):
                col_width = col.get('width', 15)
                ws.column_dimensions[get_column_letter(col_idx)].width = col_width

            ws.auto_filter.ref = ws.dimensions
            wb.save(filepath)

            self.logger.info(f"Excel rapor olusturuldu: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Excel rapor olusturma hatasi: {e}")
            return ''

    def generate_invoice_pdf(self, invoice_data: Dict, details: List[Dict]) -> str:
        if not HAS_REPORTLAB:
            return ''

        try:
            output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                'reports', 'invoices'
            )
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(
                output_dir,
                f"fatura_{invoice_data.get('InvoiceNumber', 'unknown')}.pdf"
            )

            doc = SimpleDocTemplate(filepath, pagesize=A4,
                                    rightMargin=10*mm, leftMargin=10*mm,
                                    topMargin=10*mm, bottomMargin=15*mm)
            styles = getSampleStyleSheet()
            font_name = self._get_font_name()
            story = []

            header_data = [
                [Paragraph(f"<b>FATURA</b>", ParagraphStyle('h1', fontName=font_name, fontSize=18))],
                [Paragraph(
                    f"Fatura No: {invoice_data.get('InvoiceNumber', '')}<br/>"
                    f"Tarih: {invoice_data.get('InvoiceDate', '')}<br/>"
                    f"Vade: {invoice_data.get('DueDate', '')}",
                    ParagraphStyle('hdr', fontName=font_name, fontSize=9)
                )]
            ]
            header_table = Table(header_data, colWidths=[180*mm])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2C3E50')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 5*mm))

            customer_info = [
                [Paragraph(f"<b>CARI BILGILERI</b>", ParagraphStyle('sec', fontName=font_name, fontSize=11))],
                [Paragraph(
                    f"{invoice_data.get('CurrentAccountName', '')}<br/>"
                    f"Vergi Dairesi: {invoice_data.get('TaxOffice', '')}<br/>"
                    f"Vergi No: {invoice_data.get('TaxNumber', '')}<br/>"
                    f"Adres: {invoice_data.get('Address', '')}",
                    ParagraphStyle('info', fontName=font_name, fontSize=8)
                )]
            ]
            customer_table = Table(customer_info, colWidths=[180*mm])
            customer_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D5DBDB')),
            ]))
            story.append(customer_table)
            story.append(Spacer(1, 5*mm))

            detail_headers = ['#', 'Aciklama', 'Miktar', 'Birim Fiyat', 'Iskonto', 'KDV%', 'Tutar']
            col_widths = [8*mm, 70*mm, 18*mm, 25*mm, 20*mm, 15*mm, 24*mm]
            table_data = [detail_headers]

            for i, det in enumerate(details, 1):
                table_data.append([
                    str(i),
                    det.get('Description', ''),
                    str(det.get('Quantity', 0)),
                    f"{det.get('UnitPrice', 0):,.2f}",
                    f"{det.get('DiscountAmount', 0):,.2f}",
                    f"%{det.get('VATRate', 0)}",
                    f"{det.get('TotalAmount', 0):,.2f}"
                ])

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9F9')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 5*mm))

            totals = [
                ['Ara Toplam', '', f"{invoice_data.get('SubTotal', 0):,.2f}"],
                ['KDV Toplam', '', f"{invoice_data.get('VATAmount', 0):,.2f}"],
                ['Iskonto', '', f"{invoice_data.get('DiscountAmount', 0):,.2f}"],
                ['GENEL TOPLAM', '', f"{invoice_data.get('TotalAmount', 0):,.2f}"],
            ]
            total_table = Table(totals, colWidths=[160*mm, 0, 20*mm])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2C3E50')),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, -1), (-1, -1), font_name),
                ('BOLD', (0, -1), (1, -1), True),
            ]))
            story.append(total_table)

            doc.build(story)
            self.logger.info(f"Fatura PDF olusturuldu: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Fatura PDF olusturma hatasi: {e}")
            return ''

    def generate_check_report_pdf(self, checks_data: List[Dict]) -> str:
        columns = [
            {'field': 'CheckNo', 'label': 'Cek No', 'width': 25},
            {'field': 'BankName', 'label': 'Banka', 'width': 35},
            {'field': 'Amount', 'label': 'Tutar', 'width': 25},
            {'field': 'MaturityDate', 'label': 'Vade', 'width': 25},
            {'field': 'CheckStatus', 'label': 'Durum', 'width': 25},
            {'field': 'CurrentAccountName', 'label': 'Cari', 'width': 35},
        ]
        return self.generate_pdf_report('CEK RAPORU', checks_data, columns,
                                        f"cek_raporu_{date.today().isoformat()}.pdf")

    def generate_stock_report_pdf(self, stock_data: List[Dict]) -> str:
        columns = [
            {'field': 'StockCode', 'label': 'Stok Kodu', 'width': 25},
            {'field': 'StockName', 'label': 'Stok Adi', 'width': 40},
            {'field': 'CategoryName', 'label': 'Kategori', 'width': 30},
            {'field': 'CurrentStock', 'label': 'Mevcut Stok', 'width': 20},
            {'field': 'Unit', 'label': 'Birim', 'width': 15},
            {'field': 'PurchasePrice', 'label': 'Alis Fiyati', 'width': 20},
            {'field': 'SalePrice', 'label': 'Satis Fiyati', 'width': 20},
        ]
        return self.generate_pdf_report('STOK RAPORU', stock_data, columns,
                                        f"stok_raporu_{date.today().isoformat()}.pdf")

    def generate_aging_report_pdf(self, aging_data: List[Dict]) -> str:
        columns = [
            {'field': 'CurrentAccountCode', 'label': 'Cari Kod', 'width': 20},
            {'field': 'CurrentAccountName', 'label': 'Cari Adi', 'width': 35},
            {'field': 'VadesiGelmemis', 'label': 'Vadesi Gelmemis', 'width': 25},
            {'field': 'Gun1_30', 'label': '1-30 Gun', 'width': 20},
            {'field': 'Gun31_60', 'label': '31-60 Gun', 'width': 20},
            {'field': 'Gun61_90', 'label': '61-90 Gun', 'width': 20},
            {'field': 'Gun90Plus', 'label': '90+ Gun', 'width': 20},
            {'field': 'TotalBalance', 'label': 'Toplam Bakiye', 'width': 25},
        ]
        return self.generate_pdf_report('CARi YASLANDIRMA RAPORU', aging_data, columns,
                                        f"yaslandirma_{date.today().isoformat()}.pdf")

    def generate_balance_sheet_pdf(self, data: List[Dict]) -> str:
        columns = [
            {'field': 'AccountCode', 'label': 'Hesap Kodu', 'width': 25},
            {'field': 'AccountName', 'label': 'Hesap Adi', 'width': 50},
            {'field': 'AccountGroup', 'label': 'Grup', 'width': 35},
            {'field': 'Balance', 'label': 'Bakiye', 'width': 30},
        ]
        return self.generate_pdf_report('BILANCO', data, columns,
                                        f"bilanco_{date.today().isoformat()}.pdf")
